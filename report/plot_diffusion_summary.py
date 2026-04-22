"""Create summary tables and manuscript-ready plots from diffusion_results.xlsx.

The workbook is parsed with ``openpyxl`` in read-only mode so the script stays
compact and easy to adapt when the spreadsheet layout evolves.
"""

from __future__ import annotations

import argparse
import csv
import os
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

CACHE_ROOT = Path(__file__).resolve().parent / ".cache"
CACHE_ROOT.mkdir(parents=True, exist_ok=True)
(CACHE_ROOT / "mplconfig").mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(CACHE_ROOT / "mplconfig"))
os.environ.setdefault("XDG_CACHE_HOME", str(CACHE_ROOT))

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch, Rectangle
from matplotlib.ticker import MultipleLocator
import numpy as np

MEASUREMENT_ROWS = range(2, 10)
RATIO_ROW_GROUPS = (
    ("WAT/OXM ratio", range(13, 17)),
    ("PET/PEF ratio", range(19, 23)),
)
FF_EXPORT_ORDER = ("OPLSAA", "QMFF")
FF_PLOT_ORDER = ("QMFF", "OPLSAA")
FF_DISPLAY = {"OPLSAA": "OPLS-AA", "QMFF": "QMD-FF"}
POLYMER_EXPORT_ORDER = ("PEF", "PET")
POLYMER_PLOT_ORDER = ("PET", "PEF")
LIGAND_ORDER = ("WAT", "OXM")
LIGAND_DISPLAY = {"WAT": r"H$_2$O", "OXM": r"O$_2$"}
FF_DISPLAY_COLORS = {"OPLSAA": "#D7660D", "QMFF": "#0D79B6"}
FF_HATCHES = {"OPLSAA": "//", "QMFF": None}
RUN_MARKERS = ("o", "s", "^")
BAR_EDGE_COLOR = "#111111"
GRID_COLOR = "#8C97A3"
ERRORBAR_COLOR = "#111111"
LITERATURE_BAND_COLOR = "#D7DCE2"
LITERATURE_BAND_ALPHA = 0.70
DIFFUSION_SCALE = 1.0e9
DIFFUSION_AXIS_LABEL = r"D ($\times 10^{-9}$ cm$^2$ s$^{-1}$)"
PANEL_TITLE_SIZE = 34
AXIS_LABEL_SIZE = 27
TICK_LABEL_SIZE = 22
LEGEND_FONT_SIZE = 24
RUN_LEGEND_FONT_SIZE = 16
ANNOTATION_FONT_SIZE = 18


@dataclass(frozen=True)
class ConditionKey:
    """Identify one force-field/polymer/penetrant combination.

    Parameters
    ----------
    force_field : str
        Force-field label used in the workbook.
    polymer : str
        Polymer label used in the workbook.
    ligand : str
        Penetrant label used in the workbook.
    """

    force_field: str
    polymer: str
    ligand: str


@dataclass(frozen=True)
class Measurement:
    """Store one diffusion value and its uncertainty.

    Parameters
    ----------
    value : float
        Diffusion coefficient in cm^2/s.
    uncertainty : float
        Reported uncertainty associated with ``value``.
    uncertainty_label : str
        Label describing the uncertainty type, e.g. ``"emp sem"`` or
        ``"std"``.
    """

    value: float
    uncertainty: float
    uncertainty_label: str


@dataclass(frozen=True)
class RatioEntry:
    """Store one dimensionless ratio read from the workbook.

    Parameters
    ----------
    family : str
        Ratio family label, e.g. ``"WAT/OXM ratio"``.
    force_field : str
        Force-field label used in the workbook.
    category : str
        Context for the ratio, such as polymer or penetrant identity.
    value : float
        Ratio value.
    """

    family: str
    force_field: str
    category: str
    value: float


@dataclass(frozen=True)
class WorkbookSummary:
    """Collect diffusion and ratio summaries extracted from the workbook.

    Parameters
    ----------
    pooled : dict[ConditionKey, Measurement]
        Pooled-replica diffusion estimates with empirical SEM values.
    run_average : dict[ConditionKey, Measurement]
        Replica-average diffusion estimates with between-replica standard
        deviations.
    replicates : dict[int, dict[ConditionKey, Measurement]]
        Per-replica diffusion estimates keyed by replica index.
    ratios : list[RatioEntry]
        Ratio values stored in the workbook.
    """

    pooled: dict[ConditionKey, Measurement]
    run_average: dict[ConditionKey, Measurement]
    replicates: dict[int, dict[ConditionKey, Measurement]]
    ratios: list[RatioEntry]


@dataclass(frozen=True)
class LiteratureRange:
    """Store one literature benchmark interval used in a figure.

    Parameters
    ----------
    lower : float
        Lower bound of the benchmark interval in the plotted axis units.
    upper : float
        Upper bound of the benchmark interval in the plotted axis units.
    label : str
        Short description of the benchmark interval.
    """

    lower: float
    upper: float
    label: str


LITERATURE_DIFFUSION_RANGES = {
    ("PET", "OXM"): LiteratureRange(
        lower=3.0,
        upper=6.0,
        label="Reported room-temperature diffusion range",
    ),
    ("PEF", "OXM"): LiteratureRange(
        lower=0.3,
        upper=1.0,
        label="Reported room-temperature diffusion range",
    ),
    ("PET", "WAT"): LiteratureRange(
        lower=2.0,
        upper=3.0,
        label="Reported room-temperature diffusion range",
    ),
    ("PEF", "WAT"): LiteratureRange(
        lower=0.5,
        upper=1.0,
        label="Reported room-temperature diffusion range",
    ),
}
PET_PEF_RATIO_LITERATURE_RANGES = {
    "WAT": LiteratureRange(
        lower=3.0,
        upper=5.0,
        label="Reported PET/PEF ratio range",
    ),
    "OXM": LiteratureRange(
        lower=5.0,
        upper=10.0,
        label="Reported PET/PEF ratio range",
    ),
}


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments.

    Returns
    -------
    argparse.Namespace
        Parsed arguments with workbook and output locations.
    """

    root = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=root / "diffusion_results.xlsx",
        help="Path to the diffusion workbook.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=root / "diffusion_figures",
        help="Directory where figures and CSV summaries will be written.",
    )
    return parser.parse_args()


def _cell_value(worksheet: Worksheet, cell_ref: str) -> str | float | int | None:
    """Return the value stored in one worksheet cell.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet containing the requested cell.
    cell_ref : str
        A1-style cell reference.

    Returns
    -------
    str or float or int or None
        Cell value as returned by ``openpyxl``.
    """

    return worksheet[cell_ref].value


def _read_sheet_rows(
    worksheet: Worksheet,
    *,
    row_numbers: range,
    columns: tuple[str, ...],
) -> dict[int, dict[str, str | float | int | None]]:
    """Read selected worksheet rows into sparse dictionaries.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to read.
    row_numbers : range
        Row numbers that should be copied from the worksheet.
    columns : tuple[str, ...]
        Column letters to extract for each requested row.

    Returns
    -------
    dict[int, dict[str, str | float | int | None]]
        Mapping keyed by Excel row number, where each value contains the
        requested column-letter cells.
    """

    rows: dict[int, dict[str, str | float | int | None]] = {}
    for row_number in row_numbers:
        rows[row_number] = {
            column: _cell_value(worksheet, f"{column}{row_number}") for column in columns
        }
    return rows


def _extract_measurements(
    rows: dict[int, dict[str, str | float | int | None]],
    uncertainty_label: str,
) -> dict[ConditionKey, Measurement]:
    """Extract the main diffusion table from one worksheet.

    Parameters
    ----------
    rows : dict[int, dict[str, str | float | int | None]]
        Sparse worksheet representation keyed by Excel row number.
    uncertainty_label : str
        Label used for the uncertainty column in the output objects.

    Returns
    -------
    dict[ConditionKey, Measurement]
        Diffusion values for the eight force-field/polymer/ligand conditions.
    """

    measurements: dict[ConditionKey, Measurement] = {}
    for row_number in MEASUREMENT_ROWS:
        row = rows[row_number]
        key = ConditionKey(
            force_field=str(row["A"]),
            polymer=str(row["B"]),
            ligand=str(row["C"]),
        )
        measurements[key] = Measurement(
            value=float(row["E"]),
            uncertainty=float(row["F"]),
            uncertainty_label=uncertainty_label,
        )
    return measurements


def _extract_ratios(rows: dict[int, dict[str, str | float | int | None]]) -> list[RatioEntry]:
    """Extract ratio tables from one worksheet.

    Parameters
    ----------
    rows : dict[int, dict[str, str | float | int | None]]
        Sparse worksheet representation keyed by Excel row number.

    Returns
    -------
    list[RatioEntry]
        Ratio entries in workbook order.
    """

    ratios: list[RatioEntry] = []
    for family, row_numbers in RATIO_ROW_GROUPS:
        current_ff: str | None = None
        for row_number in row_numbers:
            row = rows[row_number]
            if row.get("A") is not None:
                current_ff = str(row["A"])
            category = str(row["D"])
            ratios.append(
                RatioEntry(
                    family=family,
                    force_field=current_ff or "",
                    category=category,
                    value=float(row["E"]),
                )
            )
    return ratios


def load_workbook_summary(workbook_path: Path) -> WorkbookSummary:
    """Parse the diffusion workbook into typed summary objects.

    Parameters
    ----------
    workbook_path : pathlib.Path
        Path to the ``.xlsx`` workbook.

    Returns
    -------
    WorkbookSummary
        Parsed diffusion measurements and ratios.
    """

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    columns = ("A", "B", "C", "D", "E", "F")
    try:
        pooled_rows = _read_sheet_rows(
            workbook["pooled"],
            row_numbers=range(1, 23),
            columns=columns,
        )
        run_average_rows = _read_sheet_rows(
            workbook["run_avg"],
            row_numbers=range(1, 23),
            columns=columns,
        )

        replicate_rows: dict[int, dict[int, dict[str, str | float | int | None]]] = {}
        for replica_index in range(3):
            replicate_rows[replica_index] = _read_sheet_rows(
                workbook[f"run_{replica_index}"],
                row_numbers=range(1, 23),
                columns=columns,
            )
    finally:
        workbook.close()

    pooled = _extract_measurements(pooled_rows, uncertainty_label="emp sem")
    run_average = _extract_measurements(run_average_rows, uncertainty_label="std")
    ratios = _extract_ratios(pooled_rows)
    replicates = {
        replica_index: _extract_measurements(rows, uncertainty_label="emp sem")
        for replica_index, rows in replicate_rows.items()
    }
    return WorkbookSummary(
        pooled=pooled,
        run_average=run_average,
        replicates=replicates,
        ratios=ratios,
    )


def write_measurement_summary(summary: WorkbookSummary, output_path: Path) -> None:
    """Write a compact CSV table with diffusion measurements.

    Parameters
    ----------
    summary : WorkbookSummary
        Parsed workbook summary.
    output_path : pathlib.Path
        Destination for the CSV file.
    """

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "dataset",
                "force_field",
                "polymer",
                "ligand",
                "value_cm2_s",
                "uncertainty_cm2_s",
                "uncertainty_label",
            ]
        )
        for dataset_name, measurements in (
            ("pooled", summary.pooled),
            ("run_avg", summary.run_average),
        ):
            for key in _ordered_keys(measurements):
                entry = measurements[key]
                writer.writerow(
                    [
                        dataset_name,
                        key.force_field,
                        key.polymer,
                        key.ligand,
                        f"{entry.value:.6e}",
                        f"{entry.uncertainty:.6e}",
                        entry.uncertainty_label,
                    ]
                )


def write_ratio_summary(summary: WorkbookSummary, output_path: Path) -> None:
    """Write a compact CSV table with ratio values.

    Parameters
    ----------
    summary : WorkbookSummary
        Parsed workbook summary.
    output_path : pathlib.Path
        Destination for the CSV file.
    """

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["family", "force_field", "category", "value"])
        for entry in summary.ratios:
            writer.writerow([entry.family, entry.force_field, entry.category, f"{entry.value:.6f}"])


def _ordered_keys(measurements: dict[ConditionKey, Measurement]) -> list[ConditionKey]:
    """Return measurement keys in manuscript-friendly order.

    Parameters
    ----------
    measurements : dict[ConditionKey, Measurement]
        Measurement mapping to order.

    Returns
    -------
    list[ConditionKey]
        Keys ordered by force field, polymer, and ligand.
    """

    order = {label: index for index, label in enumerate(FF_EXPORT_ORDER)}
    polymer_order = {label: index for index, label in enumerate(POLYMER_EXPORT_ORDER)}
    ligand_order = {label: index for index, label in enumerate(LIGAND_ORDER)}
    return sorted(
        measurements,
        key=lambda key: (
            order[key.force_field],
            polymer_order[key.polymer],
            ligand_order[key.ligand],
        ),
    )


def _build_matrix(
    measurements: dict[ConditionKey, Measurement],
    *,
    force_field_order: tuple[str, ...] = FF_PLOT_ORDER,
    polymer_order: tuple[str, ...] = POLYMER_PLOT_ORDER,
) -> tuple[np.ndarray, np.ndarray]:
    """Convert measurements into ordered value and uncertainty matrices.

    Parameters
    ----------
    measurements : dict[ConditionKey, Measurement]
        Measurement mapping to convert.

    Returns
    -------
    tuple[numpy.ndarray, numpy.ndarray]
        Arrays with shape
        ``(len(force_field_order), len(polymer_order), len(LIGAND_ORDER))`` for
        values and uncertainties.
    """

    values = np.zeros((len(force_field_order), len(polymer_order), len(LIGAND_ORDER)))
    errors = np.zeros_like(values)
    for ff_index, force_field in enumerate(force_field_order):
        for polymer_index, polymer in enumerate(polymer_order):
            for ligand_index, ligand in enumerate(LIGAND_ORDER):
                entry = measurements[ConditionKey(force_field, polymer, ligand)]
                values[ff_index, polymer_index, ligand_index] = entry.value
                errors[ff_index, polymer_index, ligand_index] = entry.uncertainty
    return values, errors


def _bar_style(force_field: str) -> dict[str, str | float]:
    """Return the Matplotlib bar style used for one force field.

    Parameters
    ----------
    force_field : str
        Force-field label used in the workbook.

    Returns
    -------
    dict[str, str | float]
        Keyword arguments passed to ``Axes.bar``.
    """

    style: dict[str, str | float] = {
        "facecolor": FF_DISPLAY_COLORS[force_field],
        "edgecolor": BAR_EDGE_COLOR,
        "linewidth": 1.8,
        "alpha": 1.0,
    }
    hatch = FF_HATCHES[force_field]
    if hatch is not None:
        style["hatch"] = hatch
    return style


def _run_marker_handles() -> list[plt.Line2D]:
    """Create legend handles for the replica markers.

    Returns
    -------
    list[matplotlib.lines.Line2D]
        Marker handles corresponding to the three replica runs.
    """

    return [
        plt.Line2D(
            [0],
            [0],
            marker=marker,
            linestyle="none",
            markerfacecolor="white",
            markeredgecolor=BAR_EDGE_COLOR,
            markeredgewidth=1.5,
            markersize=9,
            label=f"run {replica_index}",
        )
        for replica_index, marker in enumerate(RUN_MARKERS)
    ]


def _literature_handle(label: str) -> Patch:
    """Create one legend handle describing literature benchmark bands.

    Parameters
    ----------
    label : str
        Legend label to show for the benchmark band.

    Returns
    -------
    matplotlib.patches.Patch
        Patch matching the benchmark band styling used in the figures.
    """

    return Patch(
        facecolor=LITERATURE_BAND_COLOR,
        edgecolor="none",
        alpha=LITERATURE_BAND_ALPHA,
        label=label,
    )


def _add_force_field_legend(fig: plt.Figure, *, include_literature: bool = False) -> None:
    """Add a figure-level legend that matches the reference styling.

    Parameters
    ----------
    fig : matplotlib.figure.Figure
        Figure receiving the legend.
    include_literature : bool, default=False
        If ``True``, append one handle describing the literature benchmark
        bands shown in the figure.
    """

    handles = [
        Patch(label=FF_DISPLAY[force_field], **_bar_style(force_field))
        for force_field in FF_PLOT_ORDER
    ]
    if include_literature:
        handles.append(_literature_handle(label="Literature range"))
    fig.legend(
        handles=handles,
        loc="upper center",
        bbox_to_anchor=(0.5, 1.02),
        ncols=len(handles),
        frameon=False,
        fontsize=LEGEND_FONT_SIZE,
        handlelength=2.7,
        columnspacing=1.6,
    )


def _add_run_marker_legend(fig: plt.Figure) -> None:
    """Add a compact figure-level legend for replica markers.

    Parameters
    ----------
    fig : matplotlib.figure.Figure
        Figure receiving the legend.
    """

    fig.legend(
        handles=_run_marker_handles(),
        loc="upper right",
        bbox_to_anchor=(0.985, 0.94),
        ncols=1,
        frameon=False,
        fontsize=RUN_LEGEND_FONT_SIZE,
        handlelength=1.0,
        handletextpad=0.5,
        labelspacing=0.25,
        borderaxespad=0.0,
    )


def _style_axis(axis: plt.Axes, *, title: str, ylabel: str | None = None) -> None:
    """Apply the common axis styling inspired by the reference figure.

    Parameters
    ----------
    axis : matplotlib.axes.Axes
        Axis to style.
    title : str
        Panel title.
    ylabel : str or None, default=None
        Optional y-axis label.
    """

    axis.set_title(title, fontsize=PANEL_TITLE_SIZE, fontweight="bold", pad=18)
    axis.grid(axis="y", color=GRID_COLOR, linestyle="--", linewidth=1.2, alpha=0.28)
    axis.set_axisbelow(True)
    axis.tick_params(axis="both", labelsize=TICK_LABEL_SIZE, width=1.5, length=0, pad=8)
    for spine in axis.spines.values():
        spine.set_visible(True)
        spine.set_linewidth(1.6)
        spine.set_color(BAR_EDGE_COLOR)
    if ylabel is not None:
        axis.set_ylabel(ylabel, fontsize=AXIS_LABEL_SIZE, labelpad=18)


def _add_literature_band(
    axis: plt.Axes,
    *,
    x_center: float,
    band_width: float,
    literature_range: LiteratureRange,
) -> None:
    """Draw one shaded literature benchmark band behind a category.

    Parameters
    ----------
    axis : matplotlib.axes.Axes
        Axis receiving the band.
    x_center : float
        Center position of the x-axis category.
    band_width : float
        Width of the shaded band in data-space x units.
    literature_range : LiteratureRange
        Benchmark interval to render.
    """

    rectangle = Rectangle(
        (x_center - 0.5 * band_width, literature_range.lower),
        band_width,
        literature_range.upper - literature_range.lower,
        facecolor=LITERATURE_BAND_COLOR,
        edgecolor="none",
        alpha=LITERATURE_BAND_ALPHA,
        zorder=0,
    )
    axis.add_patch(rectangle)


def _annotate_bar_values(
    axis: plt.Axes,
    *,
    x_positions: np.ndarray,
    heights: np.ndarray,
    y_offset: float,
) -> None:
    """Write compact numeric labels above selected bars.

    Parameters
    ----------
    axis : matplotlib.axes.Axes
        Axis receiving the annotations.
    x_positions : numpy.ndarray
        Bar-center x coordinates.
    heights : numpy.ndarray
        Bar heights to annotate.
    y_offset : float
        Vertical offset applied above each bar height.
    """

    for x_position, height in zip(x_positions, heights, strict=True):
        axis.text(
            x_position,
            height + y_offset,
            f"{height:.2f}",
            ha="center",
            va="bottom",
            fontsize=ANNOTATION_FONT_SIZE,
            color=BAR_EDGE_COLOR,
        )


def _style_axes() -> None:
    """Apply a consistent visual style to all figures."""

    plt.rcParams.update(
        {
            "axes.spines.top": True,
            "axes.spines.right": True,
            "axes.spines.left": True,
            "axes.spines.bottom": True,
            "axes.labelsize": AXIS_LABEL_SIZE,
            "axes.titlesize": PANEL_TITLE_SIZE,
            "figure.dpi": 300,
            "font.size": 14,
            "legend.frameon": False,
            "hatch.linewidth": 1.4,
            "savefig.bbox": "tight",
        }
    )


def plot_pooled_summary(summary: WorkbookSummary, output_path: Path) -> None:
    """Create the main comparison figure using pooled diffusion estimates.

    Parameters
    ----------
    summary : WorkbookSummary
        Parsed workbook summary.
    output_path : pathlib.Path
        Path to the figure file to create.
    """

    values, errors = _build_matrix(summary.pooled)
    values *= DIFFUSION_SCALE
    errors *= DIFFUSION_SCALE
    ligands = np.arange(len(LIGAND_ORDER))
    width = 0.22
    band_width = 0.72

    fig, axes = plt.subplots(1, 2, figsize=(18.5, 8.2), sharey=True)
    max_value = float(np.max(values + errors))
    literature_max = max(
        literature_range.upper for literature_range in LITERATURE_DIFFUSION_RANGES.values()
    )
    y_limit = max(1.0, np.ceil(max(max_value, literature_max) * 1.15))
    for polymer_index, polymer in enumerate(POLYMER_PLOT_ORDER):
        axis = axes[polymer_index]
        for ligand_index, ligand in enumerate(LIGAND_ORDER):
            _add_literature_band(
                axis,
                x_center=ligands[ligand_index],
                band_width=band_width,
                literature_range=LITERATURE_DIFFUSION_RANGES[(polymer, ligand)],
            )
        for ff_index, force_field in enumerate(FF_PLOT_ORDER):
            x_positions = ligands + (ff_index - 0.5) * width
            axis.bar(
                x_positions,
                values[ff_index, polymer_index],
                width=width,
                zorder=3,
                **_bar_style(force_field),
            )
            axis.errorbar(
                x_positions,
                values[ff_index, polymer_index],
                yerr=errors[ff_index, polymer_index],
                fmt="none",
                color=ERRORBAR_COLOR,
                capsize=3,
                lw=1.8,
                capthick=1.8,
                zorder=4,
            )
        axis.set_xticks(ligands, [LIGAND_DISPLAY[ligand] for ligand in LIGAND_ORDER])
        axis.set_ylim(0.0, y_limit)
        axis.yaxis.set_major_locator(MultipleLocator(1.0))
        axis.set_xlabel("Penetrant", fontsize=AXIS_LABEL_SIZE, labelpad=12)
        _style_axis(axis, title=polymer, ylabel=DIFFUSION_AXIS_LABEL if polymer_index == 0 else None)
    _add_force_field_legend(fig, include_literature=True)
    fig.subplots_adjust(top=0.80, left=0.09, right=0.985, bottom=0.18, wspace=0.13)
    fig.savefig(output_path)
    plt.close(fig)


def plot_replica_summary(summary: WorkbookSummary, output_path: Path) -> None:
    """Create the replica-averaged comparison figure with run-level scatter.

    Parameters
    ----------
    summary : WorkbookSummary
        Parsed workbook summary.
    output_path : pathlib.Path
        Path to the figure file to create.
    """

    values, errors = _build_matrix(summary.run_average)
    values *= DIFFUSION_SCALE
    errors *= DIFFUSION_SCALE
    ligands = np.arange(len(LIGAND_ORDER))
    width = 0.22

    fig, axes = plt.subplots(1, 2, figsize=(18.5, 8.4), sharey=True)
    max_value = float(np.max(values + errors))
    y_limit = max(1.0, np.ceil(max_value * 1.18))
    for polymer_index, polymer in enumerate(POLYMER_PLOT_ORDER):
        axis = axes[polymer_index]
        for ff_index, force_field in enumerate(FF_PLOT_ORDER):
            x_positions = ligands + (ff_index - 0.5) * width
            axis.bar(
                x_positions,
                values[ff_index, polymer_index],
                width=width,
                zorder=2,
                **_bar_style(force_field),
            )
            axis.errorbar(
                x_positions,
                values[ff_index, polymer_index],
                yerr=errors[ff_index, polymer_index],
                fmt="none",
                color=ERRORBAR_COLOR,
                capsize=3,
                lw=1.8,
                capthick=1.8,
                zorder=3,
            )
            for replica_index, marker in enumerate(RUN_MARKERS):
                replica_entries = summary.replicates[replica_index]
                for ligand_index, ligand in enumerate(LIGAND_ORDER):
                    key = ConditionKey(force_field, polymer, ligand)
                    x_position = x_positions[ligand_index]
                    axis.scatter(
                        x_position,
                        replica_entries[key].value * DIFFUSION_SCALE,
                        s=70,
                        marker=marker,
                        facecolor="white",
                        edgecolor=BAR_EDGE_COLOR,
                        linewidth=1.6,
                        zorder=5,
                    )
        axis.set_xticks(ligands, [LIGAND_DISPLAY[ligand] for ligand in LIGAND_ORDER])
        axis.set_ylim(0.0, y_limit)
        axis.yaxis.set_major_locator(MultipleLocator(1.0))
        axis.set_xlabel("Penetrant", fontsize=AXIS_LABEL_SIZE, labelpad=12)
        _style_axis(axis, title=polymer, ylabel=DIFFUSION_AXIS_LABEL if polymer_index == 0 else None)
    _add_force_field_legend(fig)
    _add_run_marker_legend(fig)
    fig.subplots_adjust(top=0.80, left=0.09, right=0.985, bottom=0.18, wspace=0.13)
    fig.savefig(output_path)
    plt.close(fig)


def plot_ratio_summary(summary: WorkbookSummary, output_path: Path) -> None:
    """Create the PET/PEF suppression figure with literature reference bands.

    Parameters
    ----------
    summary : WorkbookSummary
        Parsed workbook summary.
    output_path : pathlib.Path
        Path to the figure file to create.
    """

    ratio_map = {
        (entry.family, entry.force_field, entry.category): entry.value for entry in summary.ratios
    }
    categories = np.arange(len(LIGAND_ORDER))
    width = 0.22

    fig, axis = plt.subplots(1, 1, figsize=(10.8, 8.0), sharey=False)
    for ligand_index, ligand in enumerate(LIGAND_ORDER):
        _add_literature_band(
            axis,
            x_center=categories[ligand_index],
            band_width=0.72,
            literature_range=PET_PEF_RATIO_LITERATURE_RANGES[ligand],
        )

    ratio_max = 0.0
    for ff_index, force_field in enumerate(FF_PLOT_ORDER):
        heights = np.array(
            [ratio_map[("PET/PEF ratio", force_field, ligand)] for ligand in LIGAND_ORDER]
        )
        ratio_max = max(ratio_max, float(np.max(heights)))
        x_positions = categories + (ff_index - 0.5) * width
        axis.bar(
            x_positions,
            heights,
            width=width,
            zorder=3,
            **_bar_style(force_field),
        )
        _annotate_bar_values(
            axis,
            x_positions=x_positions,
            heights=heights,
            y_offset=0.15,
        )

    literature_max = max(
        literature_range.upper for literature_range in PET_PEF_RATIO_LITERATURE_RANGES.values()
    )
    axis.axhline(1.0, color=BAR_EDGE_COLOR, lw=1.5, linestyle="--", alpha=0.8)
    axis.set_xticks(categories, [LIGAND_DISPLAY[ligand] for ligand in LIGAND_ORDER])
    axis.set_ylim(0.0, np.ceil(max(ratio_max, literature_max) * 1.10))
    axis.yaxis.set_major_locator(MultipleLocator(1.0))
    axis.set_xlabel("Penetrant", fontsize=AXIS_LABEL_SIZE, labelpad=12)
    _style_axis(
        axis,
        title="PET/PEF Diffusion Ratio",
        ylabel="PET/PEF diffusion ratio",
    )
    _add_force_field_legend(fig, include_literature=True)
    fig.subplots_adjust(top=0.80, left=0.12, right=0.985, bottom=0.18)
    fig.savefig(output_path)
    plt.close(fig)


def export_all(summary: WorkbookSummary, output_dir: Path) -> None:
    """Write CSV summaries and all plot files.

    Parameters
    ----------
    summary : WorkbookSummary
        Parsed workbook summary.
    output_dir : pathlib.Path
        Directory where outputs will be created.
    """

    output_dir.mkdir(parents=True, exist_ok=True)
    write_measurement_summary(summary, output_dir / "diffusion_measurements.csv")
    write_ratio_summary(summary, output_dir / "diffusion_ratios.csv")
    for stem, plotter in (
        ("diffusion_pooled_summary", plot_pooled_summary),
        ("diffusion_replica_summary", plot_replica_summary),
        ("diffusion_ratio_summary", plot_ratio_summary),
    ):
        for extension in ("png", "pdf"):
            plotter(summary, output_dir / f"{stem}.{extension}")


def main() -> None:
    """Run the workbook parsing and export workflow."""

    args = parse_args()
    _style_axes()
    summary = load_workbook_summary(args.workbook)
    export_all(summary, args.output_dir)


if __name__ == "__main__":
    main()
